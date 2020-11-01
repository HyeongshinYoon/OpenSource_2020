from openpyxl import Workbook, load_workbook
import cv2


def image_crop_process(n):
    face_cascade = cv2.CascadeClassifier('./xml/haarcascade_frontface.xml')
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws['A1'] = 'num'
    write_ws['B1'] = 'address'
    write_ws['C1'] = 'x'
    write_ws['D1'] = 'y'
    write_ws['E1'] = 'w'
    write_ws['F1'] = 'h'

    for i in range(n):
        address = "./image/image" + str(i+1) + ".jpg"
        print(address)
        img = cv2.imread(address)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        (x, y, w, h) = faces[0]
        resize_img = cv2.resize(
            img, None, fx=256.0/w, fy=256.0/h, interpolation=cv2.INTER_AREA)
        resize_gray = cv2.cvtColor(resize_img, cv2.COLOR_BGR2GRAY)
        resize_faces = face_cascade.detectMultiScale(resize_gray, 1.3, 5)

        (rx, ry, rw, rh) = resize_faces[0]
        y1, y2 = int(ry - int(rh/8)), int(ry + rh + int(rh/8))
        x1, x2 = int(rx - int(rw/8)), int(rx + rw + int(rw/8))

        resize_cropped = resize_img[y1:y2, x1:x2]
        print(rw)
        height, width, _ = resize_img.shape
        cv2.imwrite('crop_image{}.jpg'.format(i+1), resize_cropped)
        cv2.imwrite('resize_image{}.jpg'.format(i+1), resize_img)
        write_ws.append(
            [i+1, './resize_image{}.jpg'.format(i+1), rx, ry, rw, rh, width, height])
    write_wb.save('./image_crop.xlsx')


def image_merge_process(body, face):
    load_wb = load_workbook("./image_crop.xlsx", data_only=True)
    load_ws = load_wb['Sheet']
    row = load_ws[body+1]
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    [num, address, rx, ry, rw, rh, w, h] = row_value
    print(rx)
    gan_address = "./crop_image" + str(face) + ".jpg"
    img = cv2.imread(address)
    y1, y2 = int(ry - int(rh/8)), int(ry + rh + int(rh/8))
    x1, x2 = int(rx - int(rw/8)), int(rx + rw + int(rw/8))
    gan_img = cv2.resize(cv2.imread(gan_address),
                         (y2-y1, x2-x1), interpolation=cv2.INTER_CUBIC)

    # If you want to use only changed image using GAN, you use next line only.
    #        img[y1:y2, x1:x2] = gan_img

    # Mixing background and GAN image.
    alpha_s = 0.5
    alpha_l = 1.0 - alpha_s
    for color in range(0, 3):
        img[y1:y2, x1:x2, color] = (alpha_s * gan_img[:, :, color] +
                                    alpha_l * img[y1:y2, x1:x2, color])

    # Save mixing image.
    resize_img = cv2.resize(
        img, (w, h), interpolation=cv2.INTER_CUBIC + cv2.INTER_LINEAR)
    cv2.imwrite("merge_image{}.jpg".format(num), resize_img)


def main():
    image_crop_process(5)

    body = int(input())
    face = int(input())
    image_merge_process(body, face)


if __name__ == "__main__":
    main()
